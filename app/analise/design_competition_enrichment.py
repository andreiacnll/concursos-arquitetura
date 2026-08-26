from __future__ import annotations

import json
import os
import re
import zipfile
from pathlib import Path
from typing import Any, Iterable
from xml.etree import ElementTree as ET


NUMBER_WORDS = {
    "zero": 0,
    "um": 1,
    "uma": 1,
    "dois": 2,
    "duas": 2,
    "tres": 3,
    "quatro": 4,
    "cinco": 5,
    "seis": 6,
    "sete": 7,
    "oito": 8,
    "nove": 9,
    "dez": 10,
    "onze": 11,
    "doze": 12,
    "treze": 13,
    "catorze": 14,
    "quatorze": 14,
    "quinze": 15,
    "dezasseis": 16,
    "dezassete": 17,
    "dezoito": 18,
    "dezanove": 19,
    "vinte": 20,
}


def _norm(value: object) -> str:
    text = str(value or "").lower()
    for source, target in (
        ("á", "a"), ("à", "a"), ("ã", "a"), ("â", "a"),
        ("é", "e"), ("ê", "e"), ("í", "i"),
        ("ó", "o"), ("ô", "o"), ("õ", "o"),
        ("ú", "u"), ("ç", "c"),
    ):
        text = text.replace(source, target)
    return re.sub(r"\s+", " ", text).strip()


def _clean(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def _parse_decimal(value: object) -> float | None:
    text = _clean(value).replace("\xa0", " ")
    if not text:
        return None

    match = re.search(
        r"[-+]?\d{1,3}(?:[.\s]\d{3})+(?:,\d+)?"
        r"|[-+]?\d+(?:[.,]\d+)?",
        text,
    )
    if not match:
        return None

    number = match.group(0).replace(" ", "")
    if "," in number:
        number = number.replace(".", "").replace(",", ".")
    elif number.count(".") > 1:
        number = number.replace(".", "")

    try:
        return float(number)
    except ValueError:
        return None


def _format_area(value: float | None) -> str:
    if value is None:
        return ""
    rendered = f"{value:,.2f}".replace(",", "X").replace(".", ",").replace("X", " ")
    if rendered.endswith(",00"):
        rendered = rendered[:-3]
    return f"{rendered} m²"


def _number_token(token: str) -> int | None:
    normalized = _norm(token).strip("()[]{}.,;:")
    if normalized.isdigit():
        value = int(normalized)
        return value if 0 <= value <= 100 else None
    return NUMBER_WORDS.get(normalized)


PANEL_NUMBER = (
    r"(?:\d{1,2}|zero|um|uma|dois|duas|tres|quatro|cinco|seis|"
    r"sete|oito|nove|dez|onze|doze|treze|catorze|quatorze|quinze|"
    r"dezasseis|dezassete|dezoito|dezanove|vinte)"
)
PANEL_MIXED = rf"(?:{PANEL_NUMBER}(?:\s*\(\s*{PANEL_NUMBER}\s*\))?)"


def compact_excerpt(value: object, limit: int = 420) -> str:
    text = _clean(value)
    return text if len(text) <= limit else text[:limit].rstrip() + "…"


def _panel_value(group: str) -> int | None:
    tokens = re.findall(PANEL_NUMBER, group, re.I)
    parsed = [_number_token(token) for token in tokens]
    parsed = [value for value in parsed if value is not None]
    if not parsed or len(set(parsed)) > 1:
        return None
    value = parsed[0]
    return value if 1 <= value <= 20 else None


def _panel_heading_like(text: str, match_start: int, quantity: int) -> bool:
    line_start = text.rfind("\n", 0, match_start) + 1
    line_end = text.find("\n", match_start)
    if line_end < 0:
        line_end = len(text)
    line = _norm(text[line_start:line_end])
    local = _norm(text[max(0, match_start - 45): min(len(text), match_start + 90)])
    if not line:
        return False

    # Numbered headings, including PDF-flattened ``1.1 Painéis A1`` ->
    # ``1 1 paineis a1`` after normalization. Only treat the beginning of
    # the line as a heading; later matches may be the actual requirement.
    relative_start = match_start - line_start
    if relative_start <= 10 and re.match(
        r"^\d+(?:\s*[.]\s*\d+){1,3}\s+paineis?\b", line
    ):
        return True
    if relative_start <= 10 and re.match(
        r"^\d+\s+\d+\s+paineis?\s+(?:din\s+)?a1(?:\s|$)", line
    ):
        return True
    if quantity == 1 and re.search(
        r"(?:^|\s)\d+\s+1\s+paineis?\s+(?:din\s+)?a1(?:\s|$)",
        local,
    ):
        return True
    if quantity == 1 and re.match(
        r"^(?:artigo|capitulo|secao|ponto)?\s*\d+(?:[.\s]+\d+)*\s+"
        r"paineis?\s+(?:din\s+)?a1\s*$",
        line,
    ):
        return True
    return False


def _panel_numbering_confirmation(text: str, quantity: int) -> bool:
    if quantity < 2:
        return False
    normalized = _norm(text)
    return all(
        re.search(rf"\b{index}\s*[-/]\s*{quantity}\b", normalized)
        for index in range(1, quantity + 1)
    )


def _panel_candidates(text: str, source_document: str) -> list[dict[str, Any]]:
    normalized = _norm(text)
    patterns = [
        (
            "submission_sentence",
            rf"\b(?:apresentad[ao]s?|entregues?|devem|devera|deverao|"
            rf"apresentar|entregar|concretizad[ao]s?)\b.{{0,180}}?"
            rf"\b(?P<q>{PANEL_MIXED})\s+paineis?\b",
            145,
        ),
        (
            "panel_set",
            rf"\b(?:conjunto|entrega|proposta|pecas?\s+graficas?)\b.{{0,150}}?"
            rf"\b(?:constituido|composto|formado)?\b.{{0,70}}?"
            rf"\b(?:por|de|sobre)\s+(?P<q>{PANEL_MIXED})\s+paineis?\b",
            135,
        ),
        (
            "panels_in_number_of",
            rf"\bpaineis?\s*[:\-–—,]?\s*(?:em\s+numero\s+de\s+)"
            rf"(?P<q>{PANEL_MIXED})\b",
            125,
        ),
        (
            "number_of_panels",
            rf"\b(?:numero|quantidade)\s+de\s+paineis?\s*[:\-–—]?\s*"
            rf"(?P<q>{PANEL_MIXED})\b",
            120,
        ),
        (
            "explicit_mixed_quantity",
            rf"\b(?P<q>{PANEL_MIXED})\s+paineis?\b",
            80,
        ),
    ]

    candidates: list[dict[str, Any]] = []
    seen: set[tuple[int, int, int]] = set()
    for method, pattern, base_score in patterns:
        for match in re.finditer(pattern, normalized, re.I | re.S):
            quantity = _panel_value(match.group("q"))
            if quantity is None or _panel_heading_like(normalized, match.start(), quantity):
                continue
            signature = (quantity, match.start(), match.end())
            if signature in seen:
                continue
            seen.add(signature)

            left = max(0, match.start() - 180)
            right = min(len(normalized), match.end() + 300)
            excerpt = compact_excerpt(normalized[left:right], 520)
            score = base_score
            mixed_tokens = re.findall(PANEL_NUMBER, match.group("q"), re.I)
            if len(mixed_tokens) >= 2:
                score += 28
            if re.search(r"\b(?:din\s+)?a1\b", excerpt):
                score += 18
            if "vertical" in excerpt or "horizontal" in excerpt:
                score += 10
            numbering = _panel_numbering_confirmation(normalized, quantity)
            if numbering:
                score += 35
            if quantity == 1 and len(mixed_tokens) == 1:
                score -= 45

            candidates.append(
                {
                    "quantity": quantity,
                    "score": score,
                    "source_document": source_document,
                    "evidence_excerpt": excerpt,
                    "extraction_method": (
                        f"{method}+panel_numbering"
                        if numbering
                        else method
                    ),
                }
            )
    return candidates


def _panel_format_evidence(documents) -> dict[str, Any]:
    candidates: list[dict[str, Any]] = []
    for filename, raw, normalized in documents:
        source = normalized or _norm(raw)
        for match in re.finditer(r".{0,220}paineis?.{0,300}", source, re.I | re.S):
            context = compact_excerpt(match.group(0), 520)
            norm = _norm(context)
            if not re.search(r"\b(?:din\s+)?a1\b", norm):
                continue
            score = 40
            if re.search(r"\bformato\s+(?:din\s+)?a1\b", norm):
                score += 15
            orientation = (
                "vertical" if "vertical" in norm
                else "horizontal" if "horizontal" in norm
                else ""
            )
            if orientation:
                score += 12
            candidates.append(
                {
                    "score": score,
                    "format": "A1",
                    "orientation": orientation,
                    "source_document": filename,
                    "evidence_excerpt": context,
                }
            )
    if not candidates:
        return {
            "format": "",
            "orientation": "",
            "source_document": "",
            "evidence_excerpt": "",
        }
    return max(candidates, key=lambda item: (item["score"], len(item["evidence_excerpt"])))


def extract_panel_evidence(documents) -> dict[str, Any]:
    candidates: list[dict[str, Any]] = []
    for filename, raw, normalized in documents:
        candidates.extend(_panel_candidates(raw or normalized, filename))

    best = max(
        candidates,
        key=lambda item: (item["score"], item["quantity"], len(item["evidence_excerpt"])),
        default=None,
    )
    format_evidence = _panel_format_evidence(documents)
    if best is None:
        return {
            "quantity": None,
            "quantity_confirmed": False,
            "format": format_evidence.get("format", ""),
            "orientation": format_evidence.get("orientation", ""),
            "source_document": format_evidence.get("source_document", ""),
            "evidence_excerpt": format_evidence.get("evidence_excerpt", ""),
            "extraction_method": "panel_format_only" if format_evidence.get("format") else "not_found",
        }

    evidence = best["evidence_excerpt"]
    format_excerpt = format_evidence.get("evidence_excerpt", "")
    evidence_norm = _norm(evidence)
    needs_format_support = (
        bool(format_evidence.get("format"))
        and "a1" not in evidence_norm
    ) or (
        bool(format_evidence.get("orientation"))
        and _norm(format_evidence.get("orientation")) not in evidence_norm
    )
    if format_excerpt and needs_format_support:
        evidence = compact_excerpt(f"{evidence} | {format_excerpt}", 700)
    return {
        "quantity": best["quantity"],
        "quantity_confirmed": True,
        "format": format_evidence.get("format", ""),
        "orientation": format_evidence.get("orientation", ""),
        "source_document": best.get("source_document") or format_evidence.get("source_document", ""),
        "evidence_excerpt": evidence,
        "extraction_method": best["extraction_method"],
    }


def _explicit_panel_quantity(text: str) -> int | None:
    """Compatibility helper used by tests and older callers."""
    evidence = extract_panel_evidence([("", text, _norm(text))])
    return evidence["quantity"] if evidence["quantity_confirmed"] else None

def _first(patterns: list[str], text: str) -> str:
    for pattern in patterns:
        match = re.search(pattern, text, re.I | re.S)
        if match:
            return _clean(match.group(1))
    return ""


def _build_submission(documents) -> dict[str, Any]:
    joined = "\n".join(normalized for _filename, _raw, normalized in documents)
    panel_evidence = extract_panel_evidence(documents)

    booklet = bool(
        re.search(
            r"caderno\s+(?:digital\s+)?a3|caderno\s+em\s+formato\s+a3",
            joined,
        )
    )
    max_pages = _first(
        [
            r"caderno.{0,260}?maximo\s+de\s+(\d+)\s+paginas",
            r"caderno.{0,260}?ate\s+(\d+)\s+paginas",
            r"maximo\s+de\s+(\d+)\s+paginas.{0,180}?caderno",
        ],
        joined,
    )
    booklet_context = _first([r"(.{0,180}caderno.{0,520})"], joined)
    orientation = (
        "horizontal" if "horizontal" in booklet_context
        else "vertical" if "vertical" in booklet_context
        else ""
    )
    pdf = bool(re.search(r"\bpdf\b", booklet_context))
    memory = bool(re.search(r"memoria\s+descritiva", joined))
    integrated = bool(
        re.search(
            r"caderno.{0,520}?memoria\s+descritiva|"
            r"memoria\s+descritiva.{0,520}?caderno",
            joined,
            re.I | re.S,
        )
    )

    anonymity = (
        "Obrigatório"
        if "anonim" in joined and "nao decorre sob anonimato" not in joined
        else "Não exigido"
        if "nao decorre sob anonimato" in joined
        else ""
    )

    platform = ""
    for label, pattern in (
        ("acinGov", r"\bacingov\b|\bacin\b"),
        ("VORTAL", r"\bvortal\b"),
        ("Saphety", r"\bsaphety\b"),
        ("AnoGov", r"\banogov\b"),
        ("eDelivery", r"\bedelivery\b"),
    ):
        if re.search(pattern, joined):
            platform = label
            break
    if not platform and "plataforma eletronica" in joined:
        platform = "Plataforma eletrónica"

    chapters = []
    for label, pattern in (
        ("Conceito geral", r"conceito geral"),
        ("Sustentabilidade", r"sustentabilidade"),
        ("Acessibilidade e espaços exteriores", r"acessibilidade.{0,80}?espacos exteriores"),
        ("Organização interna", r"organizacao interna"),
        ("Cumprimento do programa funcional", r"cumprimento do programa funcional"),
        ("Estimativa orçamental", r"estimativa orcamental"),
        ("Representações tridimensionais", r"representacoes tridimensionais|imagens 3d"),
        ("Organogramas e diagramas", r"organogramas|diagramas"),
        ("Cortes e fachadas", r"cortes?.{0,40}?fachadas?|fachadas?.{0,40}?cortes?"),
    ):
        if re.search(pattern, joined, re.I | re.S):
            chapters.append(label)

    return {
        "physical_panels": panel_evidence,
        "digital_booklet": {
            "required": booklet,
            "name": "Caderno A3 digital" if booklet else "",
            "format": "PDF" if pdf else "",
            "page_size": "A3" if booklet else "",
            "orientation": orientation,
            "max_pages": int(max_pages) if max_pages.isdigit() else None,
        },
        "descriptive_memory": {
            "required": memory,
            "integrated_in": "Caderno A3 digital" if integrated else "",
            "chapters": chapters,
        },
        "anonymity": anonymity,
        "platform": platform,
    }

def _build_financial(documents, facts) -> dict[str, Any]:
    joined = "\n".join(
        normalized for _filename, _raw, normalized in documents
    )
    services = str(
        (facts.get("design_services_value") or {}).get("value") or ""
    )
    vat = bool(
        re.search(
            r"(?:preco contratual|servicos|honorarios).{0,340}?"
            r"(?:acresce|acrescido|mais).{0,100}?iva"
            r"|iva.{0,100}?(?:taxa legal|em vigor)",
            joined,
            re.I | re.S,
        )
    )
    return {
        "design_services_value_net": services,
        "design_services_vat_status": "excluded" if vat else "unknown",
        "design_services_vat_text": (
            "+ IVA à taxa legal em vigor" if vat else ""
        ),
        "design_services_value_display": (
            f"{services} · + IVA" if services and vat else services
        ),
    }


def _build_contract(facts) -> dict[str, Any]:
    phases = _clean((facts.get("project_phases") or {}).get("value"))
    specialties = _clean((facts.get("specialties") or {}).get("value"))
    payments = _clean((facts.get("payment_conditions") or {}).get("value"))
    phase_items = [item for item in phases.split(" · ") if item]
    specialty_items = [item for item in specialties.split(" · ") if item]
    payment_items = [item for item in payments.split(" · ") if item]
    return {
        "phases": phase_items,
        "phase_count": len(phase_items),
        "specialties": specialty_items,
        "specialty_count": len(specialty_items),
        "payment_conditions": payment_items,
        "payment_summary": (
            f"{len(payment_items)} fases de pagamento"
            if payment_items else ""
        ),
    }


def _valid_space_label(value: str) -> bool:
    normalized = _norm(value)
    if len(normalized) < 3 or len(normalized) > 150:
        return False
    if not re.search(r"[a-z]", normalized):
        return False
    forbidden = (
        "din a", "pagina", "escala", "artigo ", "capitulo ", "preco",
        "prazo", "margem", "indice", "sumario", "n de ficha",
        "pontuacoes", "ponderacoes", "estado de conservacao", "anomalias",
        "data date", "desenhador", "folha ", "total das ponderacoes",
    )
    if any(token in normalized for token in forbidden):
        return False
    if normalized in {
        "espaco", "espacos", "designacao", "quantidade", "unidades",
        "area", "area util", "area unitaria", "area total", "total",
        "subtotal", "grupo funcional", "observacoes",
    }:
        return False
    return True


def _architectural_label(value: str) -> bool:
    normalized = _norm(value)
    if not _valid_space_label(value):
        return False
    architectural = (
        "sala", "gabinete", "laboratorio", "biblioteca", "auditorio",
        "refeitorio", "cozinha", "arrumo", "instalacao sanitaria",
        "balneario", "vestiario", "secretaria", "administracao", "direcao",
        "portaria", "rececao", "arquivo", "oficina", "atelier", "ginasio",
        "pavilhao", "circulacao", "recreio", "patio", "espaco exterior",
        "area exterior", "centro de recursos", "preparacao", "apoio",
        "professores", "alunos", "multifuncoes", "polivalente", "bar",
        "cantina", "cafetaria", "economato", "arrecadacao", "vestibulo",
        "atrio", "corredor", "escada", "elevador", "estacionamento",
        "logradouro", "jardim", "campo", "desportivo", "tecnico", "tecnica",
        "implantacao", "construcao", "edificio", "coberto", "portao",
    )
    return any(token in normalized for token in architectural)


def _infer_functional_group(label: str, current_group: str = "") -> tuple[str, bool]:
    if _clean(current_group):
        return _clean(current_group), False
    normalized = _norm(label)
    groups = [
        ("Espaços pedagógicos", ("sala", "laboratorio", "oficina", "atelier", "biblioteca", "centro de recursos")),
        ("Administração e gestão", ("administracao", "direcao", "secretaria", "gabinete", "arquivo", "economato")),
        ("Espaços sociais e apoio", ("refeitorio", "cozinha", "bar", "cafetaria", "cantina", "polivalente", "auditorio", "apoio")),
        ("Desporto", ("ginasio", "pavilhao", "campo", "balneario", "vestiario", "desportivo")),
        ("Circulações e acessos", ("circulacao", "corredor", "atrio", "vestibulo", "escada", "elevador", "portaria", "portao")),
        ("Instalações técnicas", ("tecnico", "tecnica", "avac", "eletrico", "instalacao")),
        ("Espaços exteriores", ("exterior", "recreio", "patio", "logradouro", "jardim", "estacionamento", "permeavel", "impermeavel")),
    ]
    for group, aliases in groups:
        if any(alias in normalized for alias in aliases):
            return group, True
    return "Outros espaços", True


def _schedule_row_type(label: str) -> str:
    normalized = _norm(label)
    if re.search(
        r"\btotal\s+geral\b|\btotal\s+do\s+programa\b|^total$"
        r"|\barea\s+(?:util|bruta|de\s+intervencao|total)\s+total\b"
        r"|\btotal\s+das?\s+areas?\s+uteis?\b",
        normalized,
    ):
        return "total"
    if "subtotal" in normalized or re.match(r"^total\s+(?:do|da|de|dos|das)\b", normalized):
        return "subtotal"
    return "normal"


def _row_payload(
    *,
    label: str,
    quantity: int | None,
    unit: float | None,
    total: float | None,
    filename: str,
    method: str,
    confidence: float,
    group: str = "",
    page: int | None = None,
    sheet: str = "",
    row_number: int | None = None,
    row_type: str | None = None,
) -> dict[str, Any] | None:
    label = _clean(label).strip("–—-:;,. ")
    kind = row_type or _schedule_row_type(label)
    if kind == "normal":
        label_is_valid = (
            _valid_space_label(label)
            if method == "xlsx_table"
            else _architectural_label(label)
        )
        if not label_is_valid:
            return None
    if kind != "normal" and not _valid_space_label(label):
        return None
    if total is None or total <= 0:
        return None
    if kind == "normal":
        if quantity is None or quantity <= 0 or unit is None or unit <= 0:
            return None
        expected = quantity * unit
        if abs(expected - total) > max(1.0, expected * 0.05):
            return None
    functional_group, inferred = _infer_functional_group(label, group)
    value = (
        f"{quantity} × {_format_area(unit)} = {_format_area(total)}"
        if quantity and unit is not None
        else _format_area(total)
    )
    return {
        "label": label,
        "value": value,
        "kind": "functional_area" if kind == "normal" else "schedule_total",
        "row_type": kind,
        "quantity": quantity,
        "unit_area_m2": unit,
        "total_area_m2": total,
        "functional_group": functional_group,
        "functional_group_inferred": inferred,
        "source_document": filename,
        "page": page,
        "sheet": sheet,
        "source_row": row_number,
        "confidence": confidence,
        "reconstruction_method": method,
    }


def _row_from_line(
    line: str,
    filename: str,
    *,
    method: str = "text_regex",
    page: int | None = None,
    group: str = "",
) -> dict[str, Any] | None:
    cleaned = _clean(line)
    patterns = [
        (
            re.match(
                r"^(.{3,150}?)\s+(\d{1,3})\s+"
                r"(\d{1,7}(?:[.,]\d{1,2})?)\s+"
                r"(\d{1,8}(?:[.,]\d{1,2})?)\s*(?:m2|m²)?\s*$",
                cleaned,
                re.I,
            ),
            "label_first",
        ),
        (
            re.match(
                r"^(\d{1,3})\s+(\d{1,7}(?:[.,]\d{1,2})?)\s+"
                r"(\d{1,8}(?:[.,]\d{1,2})?)\s+(.{3,150}?)\s*$",
                cleaned,
                re.I,
            ),
            "numbers_first",
        ),
    ]
    for match, order in patterns:
        if not match:
            continue
        if order == "label_first":
            label, qty_raw, unit_raw, total_raw = match.groups()
        else:
            qty_raw, unit_raw, total_raw, label = match.groups()
        quantity = int(qty_raw)
        unit = _parse_decimal(unit_raw)
        total = _parse_decimal(total_raw)
        return _row_payload(
            label=label,
            quantity=quantity,
            unit=unit,
            total=total,
            filename=filename,
            method=method,
            confidence=0.97 if method in {"pdf_layout", "xlsx_table"} else 0.91,
            group=group,
            page=page,
        )

    quantity_match = re.search(
        r"\b(\d{1,3})\s*[x×]\s*(\d{1,7}(?:[.,]\d{1,2})?)\s*m(?:2|²)?",
        cleaned,
        re.I,
    )
    if not quantity_match:
        return None
    label = cleaned[:quantity_match.start()]
    quantity = int(quantity_match.group(1))
    unit = _parse_decimal(quantity_match.group(2))
    total_match = re.search(
        r"(?:=|total\s*[:\-]?)\s*(\d{1,8}(?:[.,]\d{1,2})?)\s*m(?:2|²)?",
        cleaned[quantity_match.end():],
        re.I,
    )
    total = _parse_decimal(total_match.group(1)) if total_match else None
    if total is None and unit is not None:
        total = quantity * unit
    return _row_payload(
        label=label,
        quantity=quantity,
        unit=unit,
        total=total,
        filename=filename,
        method=method,
        confidence=0.90,
        group=group,
        page=page,
    )


def _iter_strings(value: Any, depth: int = 0) -> Iterable[str]:
    if depth > 16 or value is None:
        return
    if isinstance(value, str):
        yield value
    elif isinstance(value, dict):
        for nested in value.values():
            yield from _iter_strings(nested, depth + 1)
    elif isinstance(value, (list, tuple, set)):
        for nested in value:
            yield from _iter_strings(nested, depth + 1)


def _manifest_roots(ficha: dict[str, Any] | None) -> list[Path]:
    roots: list[Path] = []
    if isinstance(ficha, dict):
        candidates = [
            ficha.get("documentos", {}).get("source_manifest", {}).get("root"),
            ficha.get("document_audit", {}).get("source_manifest", {}).get("root"),
            ficha.get("document_insights", {}).get("source_manifest", {}).get("root"),
        ]
        for value in candidates:
            if value:
                path = Path(str(value))
                if path.exists() and path not in roots:
                    roots.append(path)
    cwd = Path.cwd()
    if cwd.exists() and cwd not in roots:
        roots.append(cwd)
    return roots


def _source_path_strings(ficha: dict[str, Any] | None) -> list[str]:
    if not isinstance(ficha, dict):
        return []
    output: list[str] = []
    seen: set[str] = set()
    for text in _iter_strings(ficha):
        clean = text.strip().strip('"\'')
        lower = clean.lower()
        if not re.search(r"\.(?:pdf|xlsx|xls)$", lower):
            continue
        signature = clean.replace("\\", "/").casefold()
        if signature in seen:
            continue
        seen.add(signature)
        output.append(clean)
    return output


def _resolve_source_paths(
    ficha: dict[str, Any] | None,
    extensions: tuple[str, ...],
) -> list[tuple[Path, str]]:
    roots = _manifest_roots(ficha)
    strings = _source_path_strings(ficha)
    output: list[tuple[Path, str]] = []
    seen: set[str] = set()

    def add(path: Path, display: str) -> None:
        try:
            resolved = path.resolve()
        except Exception:
            resolved = path
        signature = str(resolved).casefold()
        if resolved.exists() and resolved.is_file() and signature not in seen:
            seen.add(signature)
            output.append((resolved, display))

    for raw in strings:
        normalized = raw.replace("\\", "/")
        if not normalized.lower().endswith(extensions):
            continue
        direct = Path(raw)
        add(direct, raw)
        parts = [part for part in normalized.split("/") if part not in {"", "."}]
        for root in roots:
            add(root.joinpath(*parts), raw)
            add(root / Path(normalized).name, raw)

    # Basename fallback, restricted to exact names and with expensive build /
    # dependency directories pruned. This makes a project-root search safe.
    wanted = {
        Path(raw.replace("\\", "/")).name.casefold(): raw
        for raw in strings
        if raw.lower().endswith(extensions)
    }
    ignored_dirs = {
        ".git", ".venv", "venv", "node_modules", ".next", "dist",
        "build", "__pycache__", ".pytest_cache", ".mypy_cache",
        "backups", "backup",
    }
    if wanted:
        for root in roots:
            if not root.exists() or root.name.casefold() in ignored_dirs:
                continue
            try:
                for dirpath, dirnames, filenames in os.walk(root):
                    dirnames[:] = [
                        name for name in dirnames
                        if name.casefold() not in ignored_dirs
                    ]
                    for filename in filenames:
                        key = filename.casefold()
                        if key in wanted:
                            add(Path(dirpath) / filename, wanted[key])
                    if len(seen) >= len(wanted):
                        break
            except (OSError, PermissionError):
                continue
    return output


def _xlsx_values_stdlib(path: Path) -> list[tuple[str, list[list[Any]]]]:
    ns = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main", "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships"}
    rel_ns = {"p": "http://schemas.openxmlformats.org/package/2006/relationships"}
    with zipfile.ZipFile(path) as archive:
        shared: list[str] = []
        if "xl/sharedStrings.xml" in archive.namelist():
            root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
            for item in root.findall("m:si", ns):
                shared.append("".join(node.text or "" for node in item.findall(".//m:t", ns)))
        workbook = ET.fromstring(archive.read("xl/workbook.xml"))
        rels = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        targets = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels.findall("p:Relationship", rel_ns)}
        output: list[tuple[str, list[list[Any]]]] = []
        for sheet in workbook.findall("m:sheets/m:sheet", ns):
            name = sheet.attrib.get("name", "Folha")
            rid = sheet.attrib.get(f"{{{ns['r']}}}id", "")
            target = targets.get(rid, "")
            if not target:
                continue
            target = target.lstrip("/")
            if not target.startswith("xl/"):
                target = "xl/" + target
            xml = ET.fromstring(archive.read(target))
            matrix: list[list[Any]] = []
            for row in xml.findall(".//m:sheetData/m:row", ns):
                values: dict[int, Any] = {}
                for cell in row.findall("m:c", ns):
                    ref = cell.attrib.get("r", "A1")
                    letters = re.match(r"[A-Z]+", ref)
                    if not letters:
                        continue
                    column = 0
                    for char in letters.group(0):
                        column = column * 26 + (ord(char) - 64)
                    column -= 1
                    cell_type = cell.attrib.get("t", "")
                    value_node = cell.find("m:v", ns)
                    inline = cell.find("m:is", ns)
                    raw: Any = ""
                    if inline is not None:
                        raw = "".join(node.text or "" for node in inline.findall(".//m:t", ns))
                    elif value_node is not None:
                        raw = value_node.text or ""
                        if cell_type == "s":
                            try:
                                raw = shared[int(raw)]
                            except (ValueError, IndexError):
                                pass
                        elif cell_type not in {"str", "inlineStr"}:
                            try:
                                raw = float(raw)
                                if raw.is_integer():
                                    raw = int(raw)
                            except ValueError:
                                pass
                    values[column] = raw
                if values:
                    width = max(values) + 1
                    matrix.append([values.get(index, "") for index in range(width)])
                else:
                    matrix.append([])
            output.append((name, matrix))
        return output


def _xlsx_sheets(path: Path) -> list[tuple[str, list[list[Any]]]]:
    try:
        from openpyxl import load_workbook  # type: ignore
        workbook = load_workbook(path, read_only=True, data_only=True)
        output = []
        for sheet in workbook.worksheets:
            output.append((sheet.title, [list(row) for row in sheet.iter_rows(values_only=True)]))
        workbook.close()
        return output
    except Exception:
        try:
            return _xlsx_values_stdlib(path)
        except Exception:
            return []


def _header_columns(row: list[Any]) -> dict[str, int]:
    normalized = [_norm(value) for value in row]
    columns: dict[str, int] = {}
    area_candidates: list[int] = []
    for index, value in enumerate(normalized):
        if not value:
            continue
        if "grupo" in value or "area funcional" in value or "setor" in value or "sector" in value:
            columns.setdefault("group", index)
        elif (
            "quant" in value
            or value in {"qtd", "qt", "n", "n o", "nº", "numero", "unidades", "un"}
            or value.startswith("n de ")
        ):
            columns.setdefault("quantity", index)
        elif "area" in value and any(token in value for token in ("unit", "por unidade", "unidade")):
            columns.setdefault("unit", index)
        elif "area" in value and any(token in value for token in ("total", "global")):
            columns.setdefault("total", index)
        elif value in {"espaco", "espacos", "designacao", "compartimento", "programa", "descricao", "local"} or "designacao" in value:
            columns.setdefault("label", index)
        elif "area" in value or value in {"m2", "m²"}:
            area_candidates.append(index)

    # Some official workbooks use two plain area columns. In that case the
    # first is unit area and the last is total area.
    unused = [index for index in area_candidates if index not in columns.values()]
    if "unit" not in columns and unused:
        columns["unit"] = unused[0]
    if "total" not in columns and len(unused) >= 2:
        columns["total"] = unused[-1]
    return columns


def _combined_header(previous: list[Any], current: list[Any]) -> list[str]:
    width = max(len(previous), len(current))
    return [
        _clean(f"{previous[index] if index < len(previous) else ''} "
               f"{current[index] if index < len(current) else ''}")
        for index in range(width)
    ]

def _cell(row: list[Any], index: int | None) -> Any:
    if index is None or index < 0 or index >= len(row):
        return ""
    return row[index]


def _structured_spreadsheet_tables(
    ficha: dict[str, Any] | None,
) -> list[dict[str, Any]]:
    if not isinstance(ficha, dict):
        return []
    output: list[dict[str, Any]] = []
    seen: set[tuple[str, str, int]] = set()

    def visit(value: Any, depth: int = 0) -> None:
        if depth > 14 or value is None:
            return
        if isinstance(value, dict):
            tables = value.get("structured_tables")
            if isinstance(tables, list):
                for table in tables:
                    if not isinstance(table, dict):
                        continue
                    signature = (
                        _norm(table.get("source_document")),
                        _norm(table.get("sheet_name")),
                        int(table.get("header_row") or 0),
                    )
                    if signature in seen:
                        continue
                    seen.add(signature)
                    output.append(table)
            for nested in value.values():
                visit(nested, depth + 1)
        elif isinstance(value, list):
            for nested in value:
                visit(nested, depth + 1)

    visit(ficha)
    return output


def _rows_from_spreadsheet_table(table: dict[str, Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for item in table.get("rows") or []:
        if not isinstance(item, dict):
            continue
        row = dict(item)
        row.setdefault("row_type", "normal")
        row.setdefault("kind", "functional_area")
        row.setdefault("source_document", table.get("source_document", ""))
        row.setdefault("sheet", table.get("sheet_name", ""))
        rows.append(row)

    for metric in table.get("global_metrics") or []:
        if not isinstance(metric, dict):
            continue
        row = dict(metric)
        row["row_type"] = "total"
        row["kind"] = "global_area"
        row["global_key"] = str(metric.get("key") or "")
        row.setdefault("source_document", table.get("source_document", ""))
        row.setdefault("sheet", table.get("sheet_name", ""))
        rows.append(row)
    return rows


def _xlsx_area_rows(ficha: dict[str, Any] | None) -> list[dict[str, Any]]:
    source_sets: list[tuple[int, list[dict[str, Any]]]] = []

    for table in _structured_spreadsheet_tables(ficha):
        if table.get("table_type") != "functional_area_schedule":
            continue
        rows = _rows_from_spreadsheet_table(table)
        reliable = sum(1 for row in rows if row.get("row_type") == "normal")
        if reliable:
            source_sets.append((reliable, rows))

    if not source_sets:
        try:
            from app.analise.reader.spreadsheet_reader import (
                read_spreadsheet_document,
            )
        except Exception:
            read_spreadsheet_document = None

        if read_spreadsheet_document is not None:
            candidates = _resolve_source_paths(ficha, (".xlsx", ".xls", ".csv"))
            ranked = sorted(
                candidates,
                key=lambda pair: (
                    -(150 if "quadro" in _norm(pair[1]) and "area" in _norm(pair[1]) else 0)
                    -(130 if "mapa" in _norm(pair[1]) and "area" in _norm(pair[1]) else 0)
                    -(80 if "programa" in _norm(pair[1]) else 0),
                    pair[1].casefold(),
                ),
            )
            for path, display in ranked[:10]:
                try:
                    result = read_spreadsheet_document(path, display_name=display)
                except Exception:
                    continue
                for table in result.get("tables") or []:
                    if not isinstance(table, dict):
                        continue
                    rows = _rows_from_spreadsheet_table(table)
                    reliable = sum(
                        1 for row in rows if row.get("row_type") == "normal"
                    )
                    if reliable:
                        source_sets.append((reliable, rows))

    if not source_sets:
        return []
    source_sets.sort(key=lambda item: item[0], reverse=True)
    return source_sets[0][1]


def _pdf_area_rows(ficha: dict[str, Any] | None) -> list[dict[str, Any]]:
    try:
        from pypdf import PdfReader
    except Exception:
        return []
    candidates = _resolve_source_paths(ficha, (".pdf",))
    ranked = sorted(
        candidates,
        key=lambda pair: (
            -(150 if "quadro" in _norm(pair[1]) and "area" in _norm(pair[1]) else 0)
            -(140 if "mapa" in _norm(pair[1]) and "area" in _norm(pair[1]) else 0)
            -(90 if "programa funcional" in _norm(pair[1]) else 0)
            -(80 if "programa preliminar" in _norm(pair[1]) else 0),
            pair[1].casefold(),
        ),
    )
    source_sets: list[tuple[int, list[dict[str, Any]]]] = []
    for path, display in ranked[:16]:
        try:
            reader = PdfReader(str(path))
        except Exception:
            continue
        dedicated = ("quadro" in _norm(display) or "mapa" in _norm(display)) and "area" in _norm(display)
        inside = dedicated
        current_group = ""
        rows: list[dict[str, Any]] = []
        for page_number, page in enumerate(reader.pages, start=1):
            try:
                layout = page.extract_text(extraction_mode="layout") or ""
            except TypeError:
                layout = page.extract_text() or ""
            except Exception:
                continue
            normalized = _norm(layout)
            if any(marker in normalized for marker in (
                "quadro de areas", "mapa de areas", "programa funcional",
                "areas uteis previstas", "area unitaria", "area total",
            )):
                inside = True
            if not inside:
                continue
            for line in layout.splitlines():
                clean_line = _clean(line)
                norm_line = _norm(clean_line)
                if not clean_line:
                    continue
                if re.match(r"^(?:grupo|area funcional|setor|sector)\s*[:\-]", norm_line):
                    current_group = re.sub(r"^[^:\-]+[:\-]\s*", "", clean_line).strip()
                    continue
                row = _row_from_line(
                    clean_line,
                    display,
                    method="pdf_layout",
                    page=page_number,
                    group=current_group,
                )
                if row:
                    rows.append(row)
        reliable = sum(1 for row in rows if row.get("row_type") == "normal")
        if reliable:
            source_sets.append((reliable, rows))
    if not source_sets:
        return []
    source_sets.sort(key=lambda item: item[0], reverse=True)
    return source_sets[0][1]


def _text_area_rows(documents) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for filename, raw, _normalized in documents:
        source_norm = _norm(filename)
        source_priority = (
            3 if ("quadro" in source_norm or "mapa" in source_norm) and "area" in source_norm
            else 2 if "programa" in source_norm
            else 1
        )
        current_group = ""
        for line in raw.splitlines():
            clean_line = _clean(line)
            if not clean_line:
                continue
            norm_line = _norm(clean_line)
            if re.match(r"^(?:grupo|area funcional|setor|sector)\s*[:\-]", norm_line):
                current_group = re.sub(r"^[^:\-]+[:\-]\s*", "", clean_line).strip()
                continue
            row = _row_from_line(
                clean_line,
                filename,
                method="flattened_text",
                group=current_group,
            )
            if row:
                row["source_priority"] = source_priority
                rows.append(row)
    return rows


def _dedupe_schedule_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    method_score = {
        "spreadsheet_repeated_rows": 500,
        "spreadsheet_program_area_column": 490,
        "spreadsheet_explicit_columns": 500,
        "spreadsheet_global_metric": 500,
        "xlsx_table": 400,
        "pdf_layout": 300,
        "flattened_text": 100,
        "text_regex": 80,
    }
    chosen: dict[tuple[Any, ...], dict[str, Any]] = {}
    order: list[tuple[Any, ...]] = []
    for row in rows:
        # Equal room names and areas can legitimately occur under different
        # programme codes (for example two independent waiting areas). Keep
        # the document location/code in the signature and only collapse the
        # same physical source row found more than once.
        location = (
            _norm(row.get("source_document")),
            _norm(row.get("sheet")),
            row.get("source_row") if isinstance(row.get("source_row"), int) else None,
            row.get("page") if isinstance(row.get("page"), int) else None,
        )
        has_location = any(value not in {None, ""} for value in location)
        signature = (
            _norm(row.get("functional_group")),
            _norm(row.get("code")),
            _norm(row.get("label")),
            _parse_decimal(row.get("unit_area_m2")),
            _parse_decimal(row.get("total_area_m2")),
            row.get("quantity") if isinstance(row.get("quantity"), int) else None,
            str(row.get("row_type") or "normal"),
            location if has_location else None,
        )
        if signature not in chosen:
            order.append(signature)
            chosen[signature] = row
            continue
        current = chosen[signature]
        current_score = method_score.get(str(current.get("reconstruction_method")), 0)
        new_score = method_score.get(str(row.get("reconstruction_method")), 0)
        if new_score > current_score:
            chosen[signature] = row
    return [chosen[signature] for signature in order]


def _schedule_source_rows(ficha: dict[str, Any] | None, documents) -> list[dict[str, Any]]:
    xlsx = _xlsx_area_rows(ficha)
    xlsx_normal = sum(1 for row in xlsx if row.get("row_type") == "normal")
    if xlsx_normal >= 5:
        return _dedupe_schedule_rows(xlsx)
    pdf = _pdf_area_rows(ficha)
    pdf_normal = sum(1 for row in pdf if row.get("row_type") == "normal")
    if pdf_normal >= 5:
        return _dedupe_schedule_rows(pdf)
    return _dedupe_schedule_rows(_text_area_rows(documents))


def _safe_global_area(key: str, item: dict[str, Any]) -> tuple[bool, str]:
    if not item:
        return False, "ausente"
    value = _parse_decimal(item.get("total_area_m2") or item.get("value"))
    excerpt = _norm(item.get("evidence_excerpt") or item.get("label") or "")
    explicit = {
        "area_total": ("area total", "total geral", "total do programa"),
        "area_bruta": ("area bruta",),
        "area_intervencao": ("area de intervencao", "area total de intervencao", "area do terreno"),
        "area_util": ("area util total", "total das areas uteis", "area util global", "area util de construcao"),
    }[key]
    room_context = any(token in excerpt for token in (
        "sala", "gabinete", "laboratorio", "arrumo", "preparacao",
        "vestiario", "instalacao sanitaria", "2 x", "2x",
    ))
    if value is None or value <= 0:
        return False, "sem valor numérico"
    strongly_global = any(token in excerpt for token in (
        "total", "global", "total das areas", "area do terreno",
        "area de construcao",
    ))
    large_explicit_useful_area = (
        key == "area_util"
        and value >= 100
        and "area util" in excerpt
    )
    if key == "area_util" and value < 20:
        return False, "valor demasiado pequeno para total útil"
    if value < 100 and not strongly_global:
        return False, "valor pequeno sem formulação total/global segura"
    if (
        room_context
        and not any(token in excerpt for token in explicit)
        and not large_explicit_useful_area
    ):
        return False, "contexto de compartimento"
    return True, ""


def _schedule_documental_globals(rows: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    output: dict[str, dict[str, Any]] = {}
    for row in rows:
        if row.get("row_type") not in {"subtotal", "total"}:
            continue
        if _norm(row.get("scope")) == "existing":
            # Do not present existing-building metrics as proposal totals.
            continue
        label = _norm(row.get("label"))
        explicit_key = str(row.get("global_key") or "")
        key = {
            "area_intervencao": "area_intervencao",
            "area_bruta_total": "area_bruta",
        }.get(explicit_key, "")
        if not key:
            if any(token in label for token in ("area util total", "total das areas uteis", "area util global")):
                key = "area_util"
            elif "area bruta" in label:
                key = "area_bruta"
            elif "area de intervencao" in label or "area total de intervencao" in label:
                key = "area_intervencao"
            elif "area total" in label or "total geral" in label or "total do programa" in label:
                key = "area_total"
        if not key:
            continue
        total = _parse_decimal(row.get("total_area_m2"))
        if total is None:
            continue
        output[key] = {
            "label": {
                "area_total": "Área total",
                "area_bruta": "Área bruta",
                "area_intervencao": "Área de intervenção",
                "area_util": "Área útil total",
            }[key],
            "value": _format_area(total),
            "kind": "global_area",
            "total_area_m2": total,
            "source_document": row.get("source_document", ""),
            "page": row.get("page"),
            "sheet": row.get("sheet", ""),
            "source_row": row.get("source_row"),
            "confidence": row.get("confidence", 0.97),
            "evidence_excerpt": row.get("evidence_excerpt") or row.get("label", ""),
            "reconstruction_method": row.get("reconstruction_method", ""),
            "documental": True,
        }
    return output


def _equipment_type(documents) -> str:
    joined = _norm(" ".join(raw for _filename, raw, _normalized in documents))
    for label, aliases in (
        ("equipamento escolar", ("escola", "escolar", "ensino")),
        ("equipamento de saúde", ("hospital", "centro de saude", "clinica")),
        ("equipamento cultural", ("museu", "biblioteca municipal", "teatro", "centro cultural")),
        ("equipamento desportivo", ("pavilhao desportivo", "complexo desportivo", "estadio")),
        ("equipamento público", ("equipamento publico",)),
    ):
        if any(alias in joined for alias in aliases):
            return label
    return ""


def _clean_base_summary(value: object) -> str:
    text = _clean(value)
    if not text:
        return ""
    if "processo concurso\\" in text.lower() or "anexo i -" in text.lower():
        return ""
    text = re.sub(r"\.{5,}", " ", text)
    sentences = re.split(r"(?<=[.!?])\s+", text)
    usable = []
    for sentence in sentences:
        normalized = _norm(sentence)
        if len(sentence) < 45 or any(token in normalized for token in (
            "artigo ", "clausula ", "processo n", "ficheiro", "pagina",
        )):
            continue
        usable.append(sentence.strip())
        if len(" ".join(usable)) >= 420:
            break
    return _clean(" ".join(usable))


def _expanded_summary(program: dict[str, Any], documents) -> str:
    parts: list[str] = []
    base = _clean_base_summary(program.get("summary"))
    intervention = _clean(program.get("intervention_type"))
    equipment = _equipment_type(documents)
    if intervention or equipment:
        subject = equipment or "equipamento"
        action = intervention.lower() if intervention else "intervenção"
        parts.append(f"A operação corresponde a uma {action} de {subject}.")
    if base:
        parts.append(base)

    rows = [row for row in (program.get("area_schedule") or {}).get("rows", []) if row.get("row_type") == "normal"]
    groups: list[str] = []
    for row in rows:
        group = _clean(row.get("functional_group"))
        if group and group not in groups:
            groups.append(group)
    if groups:
        parts.append(
            "O programa organiza-se por "
            + ", ".join(group.lower() for group in groups[:7])
            + ", articulando espaços principais, apoios e circulações."
        )

    spaces = [_clean(item) for item in program.get("main_spaces") or [] if _clean(item)]
    if spaces:
        parts.append("Entre os espaços identificados destacam-se " + ", ".join(spaces[:8]) + ".")

    requirements = [_clean(item) for item in program.get("requirements") or [] if _clean(item)]
    if requirements:
        parts.append("A proposta deverá assegurar " + ", ".join(item.lower() for item in requirements[:6]) + ".")

    constraints = [_clean(item) for item in program.get("constraints") or [] if _clean(item)]
    if constraints:
        parts.append("As condicionantes com impacto no desenho incluem " + ", ".join(item.lower() for item in constraints[:5]) + ".")

    joined = _clean(" ".join(parts))
    return joined[:1800].rstrip()


def _clean_program(
    program: dict[str, Any],
    documents,
    ficha: dict[str, Any] | None = None,
) -> dict[str, Any]:
    warnings = list(program.get("warnings") or [])
    original_areas = list(program.get("areas") or [])
    schedule_rows = _schedule_source_rows(ficha, documents)
    schedule_rows = _dedupe_schedule_rows(schedule_rows)
    normal_rows = [row for row in schedule_rows if row.get("row_type") == "normal"]
    total_rows = [row for row in schedule_rows if row.get("row_type") in {"subtotal", "total"}]

    # A reconstructed table is authoritative. If no table is available, retain
    # the older extracted global-area list for backwards compatibility; the
    # frontend still filters it out of the room schedule.
    program["areas"] = normal_rows if normal_rows else original_areas

    documentary_globals = _schedule_documental_globals(total_rows)
    for key in ("area_total", "area_bruta", "area_intervencao", "area_util"):
        if key in documentary_globals:
            program[key] = documentary_globals[key]
            continue
        current = program.get(key) or {}
        valid, reason = _safe_global_area(key, current)
        if not valid:
            if current:
                warnings.append(f"{current.get('label') or key} rejeitada: {reason}.")
            program[key] = {}

    program["total_area"] = (program.get("area_total") or {}).get("value", "")

    has_subtotals = any(row.get("row_type") == "subtotal" for row in total_rows)
    calculated = None
    if normal_rows and not has_subtotals:
        calculated = round(sum(float(row.get("total_area_m2") or 0) for row in normal_rows), 2)
    elif has_subtotals:
        warnings.append("O total calculado não foi apresentado porque o mapa contém subtotais e poderia existir dupla contagem.")

    methods = sorted({
        str(row.get("reconstruction_method") or "")
        for row in normal_rows
        if row.get("reconstruction_method")
    })
    if any(method.startswith("spreadsheet_") for method in methods):
        warning = (
            "Mapa de áreas reconstruído a partir da estrutura de linhas e colunas "
            "da folha de cálculo oficial; o total calculado não substitui um total documental."
        )
        if warning not in warnings:
            warnings.append(warning)
    sources = []
    for row in schedule_rows:
        source = _clean(row.get("source_document"))
        if source and source not in sources:
            sources.append(source)

    program["area_schedule"] = {
        "rows": normal_rows,
        "totals": total_rows,
        "row_count": len(normal_rows),
        "reliable_row_count": sum(1 for row in normal_rows if float(row.get("confidence") or 0) >= 0.90),
        "reconstruction_method": methods[0] if len(methods) == 1 else methods,
        "source_documents": sources,
        "calculated_total_m2": calculated,
        "calculated_total_is_documental": False,
        "has_subtotals": has_subtotals,
        "warnings": warnings,
    }
    program["warnings"] = warnings
    program["summary"] = _expanded_summary(program, documents)
    return program

def enrich_design_competition(
    *,
    documents,
    facts,
    program,
    ficha: dict[str, Any] | None = None,
) -> dict[str, Any]:
    cleaned = _clean_program(
        program["functional_program"],
        documents,
        ficha,
    )
    program["functional_program"] = cleaned
    return {
        "submission": _build_submission(documents),
        "financial": _build_financial(documents, facts),
        "contract": _build_contract(facts),
        "functional_program": cleaned,
    }
